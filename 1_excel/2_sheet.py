from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새 시트를 기본 이름으로 생성
ws.title = "MySheet" #시트 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB 형식으로 변경
ws1 = wb.create_sheet("YourSheet") #주어진 이름으로 시트 생성
ws2 = wb.create_sheet("NewSheet", 2) #2번째 인덱스에 생성됨

new_ws = wb["NewSheet"] #dictionary 형태로 워크시트에 접근
print(wb.sheetnames) #모든 시트 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")