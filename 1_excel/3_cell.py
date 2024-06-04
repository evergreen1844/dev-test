from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 셀에 1이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) #A1 셀의 정보
print(ws["A1"].value) #A1 셀의 값
print(ws["A10"].value) #값이없으면 none를 출력


ws.cell(row=1, column=1)

wb.save("sample.xlsx")